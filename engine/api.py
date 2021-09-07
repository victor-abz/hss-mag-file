import os

import api_model
import pandas as pd
from fastapi import FastAPI, Response, status
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

HOST = "127.0.0.1"
PORT = 7777

app = FastAPI()


@app.get("/hello/{name}")
def read_root(name: str):
    return f"hello {name}"


@app.post("/validate-hssmag/")
def open_explorer(model: api_model.PathModel, response: Response):

    try:
        # mifotra_data = pd.read_excel(
        #     model.path, sheet_name="HSS MAG", na_filter=False, engine="openpyxl"
        # )
        wb = load_workbook(model.path, read_only=True)  # open an Excel file and return a workbook
        if "HSS MAG" in wb.sheetnames:
            return JSONResponse(
                status_code=status.HTTP_201_CREATED, content={"success": True, "msg": "Valid"}
            )
        else:
            raise Exception("Sheet called HSS MAG Should Exist in uploaded file")
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_201_CREATED,
            content={"success": False, "msg": str(e)},
        )


@app.post("/validate-ids/")
def validat_id(model: api_model.PathModel, response: Response):

    try:
        # mifotra_data = pd.read_excel(
        #     model.path, sheet_name="HSS MAG", na_filter=False, engine="openpyxl"
        # )
        adfinance_ids = pd.read_csv(model.path)
        if "id_number" in adfinance_ids.columns and "num_complet_cpte" in adfinance_ids.columns:
            return JSONResponse(
                status_code=status.HTTP_201_CREATED, content={"success": True, "msg": "Valid"}
            )
        else:
            raise Exception(
                "Sheet called id_number and num_complet_cpte Should Exist in uploaded file"
            )
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_201_CREATED,
            content={"success": False, "msg": str(e)},
        )


@app.post("/process/")
def process(model: api_model.SubmitModel, response: Response):

    try:
        mifotra_data = pd.read_excel(
            model.hss_mag_file, sheet_name="HSS MAG", na_filter=False, engine="openpyxl"
        )
        adfinance_ids = pd.read_csv(model.adfinance_file)
        adfinance_ids.rename({"id_number": "IDNumber"}, axis=1, inplace=True)

        # cols = ['IDNumber']
        # Cleanup
        mifotra_data["IDNumber"] = mifotra_data["IdNumber"].str.replace(" ", "")

        # Select ID that don't match
        invalid_values = mifotra_data[mifotra_data["IDNumber"].apply(lambda x: len(str(x)) < 16)]

        # Select valid IDs
        valid_values = mifotra_data.loc[
            mifotra_data["IDNumber"].apply(lambda x: len(str(x)) == 16)
        ]

        # valid_values

        total_merge = valid_values.merge(
            adfinance_ids, how="left", on=["IDNumber"], indicator=True
        )

        merged_data = {k: v for k, v in total_merge.groupby("_merge")}
        matching_data = merged_data["both"]

        # matching_data
        grouped_by_employer = matching_data.groupby("EntityName")

        output_folder = model.output_folder

        # Save the matched data to excel
        for row, group in grouped_by_employer:
            sheet_name = "".join(c for c in row if c.isalnum() or c.isspace())
            group.to_csv(
                os.path.join(output_folder, f"{sheet_name[:30]}.csv"),
                columns=["num_complet_cpte", "Amount"],
                index=False,
            )

        # Save Invalid to Excel
        invalid_values
        invalid_values["EmployeeId"].astype(bool)
        cleaned = invalid_values[invalid_values["EmployeeId"].astype(bool)]
        cleaned.to_excel(os.path.join(output_folder, "unmatched.xlsx"))

        return JSONResponse(
            status_code=status.HTTP_201_CREATED, content={"success": True, "msg": "Valid"}
        )

    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_201_CREATED,
            content={"success": False, "msg": str(e)},
        )


if __name__ == "__main__":
    import asyncio

    import uvicorn

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    loop.run_until_complete(uvicorn.run("api:app", host=HOST, port=PORT, reload=True))
